# -*- coding: utf-8 -*-
"""Tests for ExcelReporter — output, sheets, sanitization, path traversal.

Uses ``tmp_path`` pytest fixture for all file I/O.  Mock data classes
replace real Pydantic models to keep tests independent.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from src.reports.excel_reporter import ExcelReporter
from src.utils.security import SecurityError

# ---------------------------------------------------------------------------
# Lightweight mock data (mirrors GapAnalysisResult / RiskReport shapes)
# ---------------------------------------------------------------------------


@dataclass
class MockControl:
    """Mock control object matching the Control model interface."""

    control_id: str
    title: str
    description: str = ""
    priority: str = "medium"


@dataclass
class MockGapResult:
    """Mock gap analysis result with sensible defaults."""

    framework_name: str = "NIST_CSF"
    total_controls: int = 5
    implemented_count: int = 2
    missing_controls: list[Any] = field(default_factory=list)
    compliance_percentage: float = 40.0
    analyzed_at: datetime = field(default_factory=lambda: datetime.now(tz=UTC))


@dataclass
class MockFinding:
    """Mock risk finding matching the RiskFinding model interface."""

    control_id: str
    control_name: str
    severity: str = "HIGH"
    likelihood: float = 0.8
    impact: float = 0.8
    risk_score: float = 64.0


@dataclass
class MockRiskReport:
    """Mock risk report matching the RiskReport model interface."""

    overall_risk_level: str = "HIGH"
    risk_score: float = 64.0
    findings: list[Any] = field(default_factory=list)
    recommendations: list[str] = field(default_factory=list)
    scored_at: datetime = field(default_factory=lambda: datetime.now(tz=UTC))


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def gap_result() -> MockGapResult:
    """Realistic gap analysis result with 3 missing controls."""
    return MockGapResult(
        missing_controls=[
            MockControl(
                control_id="ID.AM-1",
                title="Asset Inventory",
                description="Maintain asset inventory",
                priority="high",
            ),
            MockControl(
                control_id="PR.AC-1",
                title="Access Control",
                description="Manage access permissions",
                priority="critical",
            ),
            MockControl(
                control_id="DE.CM-1",
                title="Monitoring",
                description="Continuous monitoring",
                priority="medium",
            ),
        ],
    )


@pytest.fixture()
def risk_report() -> MockRiskReport:
    """Realistic risk report with findings and recommendations."""
    return MockRiskReport(
        findings=[
            MockFinding(
                control_id="ID.AM-1",
                control_name="Asset Inventory",
                severity="HIGH",
                risk_score=64.0,
            ),
            MockFinding(
                control_id="PR.AC-1",
                control_name="Access Control",
                severity="CRITICAL",
                risk_score=90.25,
            ),
            MockFinding(
                control_id="DE.CM-1",
                control_name="Monitoring",
                severity="MEDIUM",
                risk_score=27.0,
            ),
        ],
        recommendations=[
            "1. Prioritise implementation of control PR.AC-1 (Access Control)",
            "2. Prioritise implementation of control ID.AM-1 (Asset Inventory)",
            "3. Prioritise implementation of control DE.CM-1 (Monitoring)",
        ],
    )


@pytest.fixture()
def reporter(tmp_path: Path) -> ExcelReporter:
    """ExcelReporter with template_dir set to a temp directory."""
    return ExcelReporter(template_dir=tmp_path)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestGenerateReturnsValidPath:
    """generate() should return a .xlsx path and write a non-empty file."""

    # Verify that generate() returns a path ending in .xlsx
    def test_returns_xlsx_path(
        self,
        reporter: ExcelReporter,
        gap_result: MockGapResult,
        risk_report: MockRiskReport,
        tmp_path: Path,
    ) -> None:
        result = reporter.generate(gap_result, risk_report, tmp_path)  # type: ignore[arg-type]
        assert result.suffix == ".xlsx"

    # Verify that the output file exists and has content
    def test_output_file_exists_and_is_nonempty(
        self,
        reporter: ExcelReporter,
        gap_result: MockGapResult,
        risk_report: MockRiskReport,
        tmp_path: Path,
    ) -> None:
        result = reporter.generate(gap_result, risk_report, tmp_path)  # type: ignore[arg-type]
        assert result.exists()
        assert result.stat().st_size > 0


class TestSheetStructure:
    """Excel file must contain exactly 3 sheets with the correct names."""

    # Verify the workbook has the 3 expected sheets
    def test_has_three_sheets(
        self,
        reporter: ExcelReporter,
        gap_result: MockGapResult,
        risk_report: MockRiskReport,
        tmp_path: Path,
    ) -> None:
        result = reporter.generate(gap_result, risk_report, tmp_path)  # type: ignore[arg-type]
        wb = load_workbook(result)
        expected_names = {"Executive Summary", "Gap Analysis", "Risk Findings"}
        assert set(wb.sheetnames) == expected_names
        assert len(wb.sheetnames) == 3


class TestFormulaInjection:
    """Formula injection payloads in control names must be sanitized."""

    # Verify that =CMD|'/C calc' is not written raw into any cell
    def test_formula_injection_is_sanitized(
        self,
        risk_report: MockRiskReport,
        tmp_path: Path,
    ) -> None:
        reporter = ExcelReporter(template_dir=tmp_path)
        # Inject a formula into the control title
        malicious_gap = MockGapResult(
            missing_controls=[
                MockControl(
                    control_id="ID.AM-1",
                    title="=CMD|'/C calc'",
                    priority="high",
                ),
            ],
        )
        result = reporter.generate(malicious_gap, risk_report, tmp_path)  # type: ignore[arg-type]
        wb = load_workbook(result)
        ws = wb["Gap Analysis"]

        # Check that no cell starts with the raw = formula prefix
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell_value in row:
                if cell_value is not None:
                    assert not str(cell_value).startswith("="), (
                        f"Formula injection not sanitized: {cell_value}"
                    )


class TestPathTraversal:
    """Output path with traversal must raise SecurityError."""

    # Verify that a path traversal attempt is rejected
    def test_traversal_raises_security_error(
        self,
        reporter: ExcelReporter,
        gap_result: MockGapResult,
        risk_report: MockRiskReport,
        tmp_path: Path,
    ) -> None:
        malicious_path = tmp_path / ".." / ".." / "etc"
        with pytest.raises(SecurityError):
            reporter.generate(gap_result, risk_report, malicious_path)  # type: ignore[arg-type]


class TestCompliancePercentage:
    """Compliance percentage must appear in Executive Summary sheet."""

    # Verify the compliance percentage is written to the first sheet
    def test_compliance_percentage_in_summary(
        self,
        reporter: ExcelReporter,
        gap_result: MockGapResult,
        risk_report: MockRiskReport,
        tmp_path: Path,
    ) -> None:
        result = reporter.generate(gap_result, risk_report, tmp_path)  # type: ignore[arg-type]
        wb = load_workbook(result)
        ws = wb["Executive Summary"]

        # Collect all cell values in the sheet
        all_values = []
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if val is not None:
                    all_values.append(str(val))

        # Check that 40.0% appears somewhere in the values
        assert any("40.0" in v for v in all_values), (
            f"Compliance percentage not found in Executive Summary. Values: {all_values}"
        )


class TestGapAnalysisRowCount:
    """Gap Analysis sheet rows should match the number of missing controls."""

    # Verify the number of data rows equals the number of missing controls
    def test_row_count_matches_missing_controls(
        self,
        reporter: ExcelReporter,
        gap_result: MockGapResult,
        risk_report: MockRiskReport,
        tmp_path: Path,
    ) -> None:
        result = reporter.generate(gap_result, risk_report, tmp_path)  # type: ignore[arg-type]
        wb = load_workbook(result)
        ws = wb["Gap Analysis"]

        # Count data rows (skip header at row 1)
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Filter out completely empty rows
        non_empty_rows = [r for r in data_rows if any(v is not None for v in r)]
        assert len(non_empty_rows) == len(gap_result.missing_controls)
