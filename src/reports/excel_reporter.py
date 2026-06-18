# -*- coding: utf-8 -*-
"""Excel report generator using openpyxl.

Produces a professional compliance report as a standalone .xlsx workbook
with three sheets: Executive Summary, Gap Analysis, and Risk Findings.
All dependencies are injected via the constructor — template directory
is never hardcoded.
"""

from __future__ import annotations

import re
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.core.exceptions import ReportGenerationError
from src.core.gap_analyzer import GapAnalysisResult
from src.core.risk_scorer import RiskReport
from src.reports.base_reporter import BaseReporter
from src.utils.logger import get_logger

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Constants — severity-to-fill colour mapping for openpyxl
# ---------------------------------------------------------------------------

# Red fill for CRITICAL severity
FILL_CRITICAL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
# Orange fill for HIGH severity
FILL_HIGH = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
# Yellow fill for MEDIUM severity
FILL_MEDIUM = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# Green fill for LOW severity
FILL_LOW = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Map severity strings to their PatternFill constants
SEVERITY_FILLS: dict[str, PatternFill] = {
    "CRITICAL": FILL_CRITICAL,
    "HIGH": FILL_HIGH,
    "MEDIUM": FILL_MEDIUM,
    "LOW": FILL_LOW,
}

# Bold font for header rows and highlighted cells
BOLD_FONT = Font(bold=True)
# Large bold font for the title row
TITLE_FONT = Font(bold=True, size=16)

# Regex to strip HTML tags from cell values
_HTML_TAG_RE = re.compile(r"<[^>]+>")
# Characters that trigger formula injection in Excel
_FORMULA_PREFIXES = ("=", "+", "-", "@")


class ExcelReporter(BaseReporter):
    """Generate Excel compliance reports via openpyxl.

    Args:
        template_dir: Absolute ``Path`` to the template directory.
            Accepted for interface consistency with other reporters;
            not directly used by Excel generation.

    Raises:
        ReportGenerationError: If *template_dir* does not exist or
            fails security validation.
    """

    def __init__(self, template_dir: Path) -> None:
        # Store injected template directory for interface consistency
        self._template_dir = template_dir
        logger.debug(
            "ExcelReporter initialised",
            extra={"template_dir": str(template_dir)},
        )

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate(
        self,
        gap_result: GapAnalysisResult,
        risk_report: RiskReport,
        output_path: Path,
    ) -> Path:
        """Generate an Excel workbook and write it to *output_path*.

        Args:
            gap_result: Gap analysis output.
            risk_report: Risk scoring output.
            output_path: Directory to write the report into.

        Returns:
            Absolute ``Path`` to the generated ``.xlsx`` file.

        Raises:
            ReportGenerationError: On any workbook or I/O error.
            SecurityError: If *output_path* contains traversal.
        """
        # Validate output directory via safe_path
        validated_dir = self.validate_output_path(output_path)

        try:
            # Create a new workbook (removes the default sheet later)
            wb = Workbook()

            # Build all three sheets in order
            self._build_executive_summary(wb, gap_result, risk_report)
            self._build_gap_analysis(wb, gap_result)
            self._build_risk_findings(wb, risk_report)

            # Remove the default blank sheet created by openpyxl
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            # Construct output filename with framework name and timestamp
            framework_name = self._sanitize_cell_value(gap_result.framework_name)
            timestamp = datetime.now(tz=UTC).strftime("%Y%m%d_%H%M%S")
            filename = f"report_{framework_name}_{timestamp}.xlsx"
            report_path = validated_dir / filename

            # Save the workbook to disk
            wb.save(str(report_path))

        except (OSError, KeyError, ValueError) as exc:
            msg = f"Failed to generate Excel report: {exc}"
            raise ReportGenerationError(msg, details={"cause": str(exc)}) from exc

        logger.info(
            "Excel report generated",
            extra={"report_path": str(report_path)},
        )
        return report_path

    # ------------------------------------------------------------------
    # Private sheet builders
    # ------------------------------------------------------------------

    def _build_executive_summary(
        self,
        wb: Workbook,
        gap_result: GapAnalysisResult,
        risk_report: RiskReport,
    ) -> None:
        """Create Sheet 1 — Executive Summary with key metrics."""
        ws = wb.create_sheet(title="Executive Summary")

        # Title row with framework name and compliance percentage
        title_text = (
            f"{self._sanitize_cell_value(gap_result.framework_name)} — "
            f"Compliance: {gap_result.compliance_percentage}%"
        )
        ws["A1"] = title_text
        ws["A1"].font = TITLE_FONT

        # Metric labels in column A, values in column B
        metrics: list[tuple[str, str | int | float]] = [
            ("Total Controls", gap_result.total_controls),
            ("Implemented Controls", gap_result.implemented_count),
            ("Missing Controls", len(gap_result.missing_controls)),
            ("Compliance Percentage", f"{gap_result.compliance_percentage}%"),
            ("Overall Risk Level", self._sanitize_cell_value(risk_report.overall_risk_level)),
            ("Risk Score", risk_report.risk_score),
        ]

        # Write each metric row starting at row 3
        for row_idx, (label, value) in enumerate(metrics, start=3):
            ws.cell(row=row_idx, column=1, value=label).font = BOLD_FONT
            ws.cell(row=row_idx, column=2, value=value)

        # Color-code the risk level cell based on severity
        risk_level = risk_report.overall_risk_level.upper()
        risk_cell = ws.cell(row=7, column=2)  # row 7 = "Overall Risk Level"
        if risk_level in SEVERITY_FILLS:
            risk_cell.fill = SEVERITY_FILLS[risk_level]

        # Auto-adjust column widths for readability
        self._auto_adjust_columns(ws)

    def _build_gap_analysis(
        self,
        wb: Workbook,
        gap_result: GapAnalysisResult,
    ) -> None:
        """Create Sheet 2 — Gap Analysis with one row per missing control."""
        ws = wb.create_sheet(title="Gap Analysis")

        # Define header row
        headers = [
            "Control ID", "Control Name", "Family",
            "Severity", "Risk Score", "Recommendation",
        ]

        # Write header cells with bold font
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = BOLD_FONT
            cell.alignment = Alignment(horizontal="center")

        # Freeze the top row so headers stay visible while scrolling
        ws.freeze_panes = "A2"

        # Write one row per missing control
        for row_idx, ctrl in enumerate(gap_result.missing_controls, start=2):
            # Derive family from control ID prefix (e.g. "ID.AM-1" → "ID")
            control_id = self._sanitize_cell_value(ctrl.control_id)
            family = control_id.split(".")[0] if "." in control_id else control_id

            # Sanitize title for safe cell insertion
            title = self._sanitize_cell_value(ctrl.title)

            # Use priority as severity proxy, normalised to uppercase
            severity = self._sanitize_cell_value(
                getattr(ctrl, "priority", "medium")
            ).upper()

            ws.cell(row=row_idx, column=1, value=control_id)
            ws.cell(row=row_idx, column=2, value=title)
            ws.cell(row=row_idx, column=3, value=family)
            ws.cell(row=row_idx, column=4, value=severity)
            ws.cell(row=row_idx, column=5, value=0.0)  # Placeholder risk score
            ws.cell(
                row=row_idx, column=6,
                value=f"Implement control {control_id}",
            )

            # Apply severity-based row background color
            fill = SEVERITY_FILLS.get(severity)
            if fill:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = fill

        # Auto-adjust column widths to fit content
        self._auto_adjust_columns(ws)

    def _build_risk_findings(
        self,
        wb: Workbook,
        risk_report: RiskReport,
    ) -> None:
        """Create Sheet 3 — Risk Findings sorted by risk score descending."""
        ws = wb.create_sheet(title="Risk Findings")

        # Define header row
        headers = [
            "Control ID", "Likelihood", "Impact", "Risk Score", "Severity",
        ]

        # Write header cells with bold font
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = BOLD_FONT
            cell.alignment = Alignment(horizontal="center")

        # Sort findings by risk score — highest risk first
        sorted_findings = sorted(
            risk_report.findings,
            key=lambda f: f.risk_score,
            reverse=True,
        )

        # Write one row per finding
        for row_idx, finding in enumerate(sorted_findings, start=2):
            ws.cell(
                row=row_idx, column=1,
                value=self._sanitize_cell_value(finding.control_id),
            )
            ws.cell(row=row_idx, column=2, value=finding.likelihood)
            ws.cell(row=row_idx, column=3, value=finding.impact)
            ws.cell(row=row_idx, column=4, value=finding.risk_score)
            ws.cell(
                row=row_idx, column=5,
                value=self._sanitize_cell_value(finding.severity),
            )

            # Bold the top 3 highest-risk rows for emphasis
            if row_idx <= 4:  # rows 2, 3, 4 = top 3
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).font = BOLD_FONT

        # Auto-adjust column widths to fit content
        self._auto_adjust_columns(ws)

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _sanitize_cell_value(value: str) -> str:
        """Strip HTML tags, special chars, and formula-injection prefixes.

        Prevents Excel formula injection by prefixing dangerous values
        with a single quote, and removes HTML tags for clean cell text.

        Args:
            value: Raw string to sanitize.

        Returns:
            Safe string suitable for an Excel cell.
        """
        # Strip HTML tags from the value
        cleaned = _HTML_TAG_RE.sub("", str(value))

        # Remove non-printable / special characters except basic punctuation
        cleaned = re.sub(r"[^\w\s.,;:!?%()\-/]", "", cleaned)

        # Block formula injection — prefix with single quote if needed
        if cleaned and cleaned[0] in _FORMULA_PREFIXES:
            cleaned = f"'{cleaned}"

        return cleaned.strip()

    @staticmethod
    def _auto_adjust_columns(ws: Workbook) -> None:  # type: ignore[arg-type]
        """Set each column width to the widest cell value plus padding.

        Iterates every cell in the worksheet and finds the maximum string
        length per column, then sets column_dimensions accordingly.
        """
        for col_cells in ws.columns:
            max_length = 0
            # Get the column letter from the first cell
            column_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            # Add padding of 3 characters for visual spacing
            ws.column_dimensions[column_letter].width = max_length + 3
